# Import for the Web Bot
from botcity.web import WebBot, Browser, By
import selenium.webdriver
from selenium import webdriver
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from botcity.maestro import *
import openpyxl

# Create WebBot instance
bot = WebBot()
# Disable errors if we are not connected to Maestro
BotMaestroSDK.RAISE_NOT_CONNECTED = False

def main():
    cap = DesiredCapabilities().FIREFOX
    cap["marionette"] = True
    bot.browser = webdriver.Firefox(capabilities=cap)
    bot.browser.get('https://cidades.ibge.gov.br/')

    button = WebDriverWait(bot.browser, 10).until(
        EC.visibility_of_element_located((By.XPATH, '//button[text()="Comece a usar"]'))
    )
    bot.browser.execute_script("arguments[0].click();", button)

    aside = WebDriverWait(bot.browser, 10).until(
        EC.invisibility_of_element_located((By.CSS_SELECTOR, '.aside--municipio'))
    )

    estados_option = WebDriverWait(bot.browser, 1).until(
        EC.visibility_of_element_located((By.XPATH, '//span[text()="Estados"]'))
    )
    estados_option.click()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.cell(row=1, column=1, value="Estados")
    for i, uf in enumerate(ufs, start=2):
        worksheet.cell(row=i, column=1, value=uf)
    workbook.save("states.xlsx")


def not_found(label):
    print(f"Element not found: {label}")


if __name__ == '__main__':
    # Set the Firefox browser
    main()

    # Stop the browser
    bot.stop_browser()
